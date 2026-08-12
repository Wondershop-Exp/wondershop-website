"""
Order execution form generator — Excel (openpyxl) + PDF (reportlab).

V1 scope, per Shruti (2026-08-12): print only what we already capture on a
booking — LeadSubmitRequest fields + builder_snapshot. Anything genuinely
manual (vendor assigned, Event Ops Lead, payment mode, pinata bags/fillings,
gift-wrap instructions, event schedule, inventory to pack) prints as a
blank line / empty grid for ops to fill by hand on the printout. No
separate internal tool needed for this cut — that can follow once this is
live and being used.
"""
import io
import re
import asyncio
import logging
import urllib.parse
from typing import Optional

import catalogue_data as cat
from catalogue_data import SITE_BASE_URL

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepInFrame, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

logger = logging.getLogger(__name__)


# ─── formatting helpers ──────────────────────────────────────────────────

def _ordinal(n: int) -> str:
    if 10 <= n % 100 <= 20:
        suf = "th"
    else:
        suf = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suf}"


def _fmt_date(d) -> str:
    if not d:
        return "—"
    if isinstance(d, str):
        return d
    try:
        return f"{_ordinal(d.day)} {d.strftime('%b %Y')}"
    except Exception:
        return str(d)


def _fmt_money(v) -> str:
    if v is None:
        return "—"
    try:
        return f"Rs. {v:,.0f}"
    except Exception:
        return str(v)


def _child_age_gender(req) -> str:
    """'F5' style summary (first letter of gender + age), comma-joined for
    multiple kids. Degrades gracefully when either piece is missing."""
    ages = [a.strip() for a in (req.child_ages or "").split(",") if a.strip()]
    genders = [g.strip() for g in (req.child_genders or "").split(",") if g.strip()]
    if not ages and not genders:
        return "—"
    parts = []
    for i in range(max(len(ages), len(genders))):
        age = ages[i] if i < len(ages) else ""
        gen = genders[i] if i < len(genders) else ""
        letter = gen[:1].upper() if gen else ""
        parts.append((f"{letter}{age}".strip()) or "—")
    return ", ".join(parts) if parts else "—"


# ─── data assembly ────────────────────────────────────────────────────────

def assemble_order_form_data(req, lead_id: int, event_sales_lead: Optional[str],
                              reward_code: Optional[str] = None) -> dict:
    """Pulls together everything the order form needs straight from the
    booking payload already captured at checkout — no extra DB lookups."""
    snap = req.builder_snapshot or {}

    decor   = snap.get("decor") or {}
    host    = snap.get("host") or {}
    dj      = snap.get("dj") or {}
    photo   = snap.get("photo") or {}
    pinata  = snap.get("pinata") or {}
    einvite = snap.get("einvite") or {}
    activities = snap.get("activities") or []
    gifts      = snap.get("gifts") or []

    dj_addon_bits = []
    if getattr(req, "dj_lights_addon", False):
        dj_addon_bits.append("Lights")
    if getattr(req, "dj_smoke_machine_addon", False):
        dj_addon_bits.append("Smoke Machine")

    venue_type_address = " · ".join([v for v in [req.location_type, req.venue] if v]) or "—"
    venue_contact = " ".join([v for v in [
        req.venue_contact_name,
        f"({req.venue_contact_phone})" if req.venue_contact_phone else None,
    ] if v]) or "—"

    coupon = req.redeemed_coupon_code or getattr(req, "sales_lead_code", None) or None
    child_name = (req.child_names or "").split(",")[0].strip() or req.parent_name or "Guest"

    # ── catalogue lookups — reference photos + inclusion lists for whatever
    # was actually chosen on the booking. Anything that doesn't resolve to a
    # confident catalogue match (custom design, package-specific id, etc.)
    # is left out rather than guessed.
    decor_ref = cat.resolve_decor(decor.get("id"), decor.get("p"))
    photo_features = cat.PHOTO_TIER_FEATURES.get(photo.get("tier"), [])
    einvite_image_path = cat.resolve_einvite_image(einvite.get("id")) if einvite else None
    pinata_image_path = cat.resolve_pinata_image(pinata.get("id")) if pinata else None

    gifts_detail = []
    for g in gifts:
        ref = cat.resolve_gift(g.get("id"))
        unit = g.get("unit")
        if unit is None and ref:
            unit = ref["catalogue_unit"]
        qty = g.get("qty") or 0
        gifts_detail.append({
            "name": g.get("n") or "—",
            "qty": qty,
            "unit": unit,
            "total": (unit * qty) if (unit is not None and qty) else None,
            "image_path": ref["image_path"] if ref else None,
        })

    packaging_label = cat.PACKAGING_LABELS.get(snap.get("gift_packaging"), "None selected")
    thank_you_note = "Yes" if snap.get("gift_thank_you_note") else "No"

    return {
        "lead_id": lead_id,
        "child_name": child_name,
        "client_name": req.parent_name or "—",
        "client_phone": req.phone or "—",
        "child_names": req.child_names or "—",
        "child_age_gender": _child_age_gender(req),
        "event_date": _fmt_date(req.event_date),
        "event_date_raw": req.event_date,
        "event_time": req.event_time or "—",
        "venue_type_address": venue_type_address,
        "venue_contact": venue_contact,
        "kids_count": req.kids_count or "—",
        "theme": req.theme or "—",
        "event_sales_lead": event_sales_lead or "—",
        "billing_amount": _fmt_money(req.order_grand_total),
        "advance_paid": _fmt_money(req.order_advance),
        "pending_amount": _fmt_money(req.order_balance),
        "coupon_code": coupon or "—",
        "remarks": req.remarks or "—",
        "decor_name": decor.get("n") or "—",
        "host_tier": host.get("tier") or "—",
        "dj_tier": dj.get("tier") or "—",
        "dj_addons": ", ".join(dj_addon_bits) or "None",
        "photo_tier": photo.get("tier") or "—",
        "pinata_name": pinata.get("n") or "—",
        "einvite_name": einvite.get("n") or "—",
        "activities": [a.get("n") for a in activities if a.get("n")] or ["—"],
        "reward_code": reward_code or "—",
        "location": req.city or req.venue or "",
        "decor_image_path": decor_ref["image_path"] if decor_ref else None,
        "decor_spec": decor_ref["spec"] if decor_ref else [],
        "photo_features": photo_features,
        "einvite_image_path": einvite_image_path,
        "pinata_image_path": pinata_image_path,
        "gifts_detail": gifts_detail,
        "gift_packaging_label": packaging_label,
        "gift_thank_you_note": thank_you_note,
    }


async def fetch_order_form_images(data: dict) -> dict:
    """Downloads whatever reference images resolved during assembly (decor,
    e-invite, pinata, per-gift thumbnails) from the live site so they can be
    embedded in the xlsx/pdf. Best-effort — a failed fetch just means that
    one slot stays blank on the form, never blocks the email."""
    urls = {}
    if data.get("decor_image_path"):
        urls["decor"] = data["decor_image_path"]
    if data.get("einvite_image_path"):
        urls["einvite"] = data["einvite_image_path"]
    if data.get("pinata_image_path"):
        urls["pinata"] = data["pinata_image_path"]
    for i, g in enumerate(data.get("gifts_detail", [])):
        if g.get("image_path"):
            urls[f"gift{i}"] = g["image_path"]

    if not urls:
        return data

    import httpx

    async def _fetch(key, path):
        url = f"{SITE_BASE_URL}/img/{urllib.parse.quote(path)}"
        try:
            async with httpx.AsyncClient(timeout=8) as client:
                r = await client.get(url)
            if r.status_code == 200:
                return key, r.content
        except Exception as exc:
            logger.warning(f"order form: image fetch failed for {path} — {exc}")
        return key, None

    results = await asyncio.gather(*(_fetch(k, v) for k, v in urls.items()))
    bytes_by_key = dict(results)

    data["decor_image_bytes"] = bytes_by_key.get("decor")
    data["einvite_image_bytes"] = bytes_by_key.get("einvite")
    data["pinata_image_bytes"] = bytes_by_key.get("pinata")
    for i, g in enumerate(data.get("gifts_detail", [])):
        g["image_bytes"] = bytes_by_key.get(f"gift{i}")
    return data


def order_form_filename(data: dict, ext: str) -> str:
    """<customer name>-<date>-<location>-Birthday Order Form.<ext>"""
    raw = f"{data['child_name']}-{data['event_date']}-{data['location'] or ''}-Birthday Order Form.{ext}"
    safe = re.sub(r'[\\/:*?"<>|]+', "", raw)
    safe = re.sub(r"\s+", " ", safe).strip()
    return safe


# Rows we already have data for (label, value-key or literal).
_DETAIL_ROWS = [
    ("Client Name",        "client_name"),
    ("Client Number",      "client_phone"),
    ("Child Name(s)",      "child_names"),
    ("Age & Gender",       "child_age_gender"),
    ("Date",               "event_date"),
    ("Time",               "event_time"),
    ("Venue Type & Address", "venue_type_address"),
    ("Venue Contact",      "venue_contact"),
    ("# Kids",             "kids_count"),
    ("Theme",              "theme"),
    ("Event Sales Lead",   "event_sales_lead"),
]

_SERVICE_ROWS = [
    ("Host",          "host_tier"),
    ("Music (DJ)",    "dj_tier"),
    ("DJ Add-ons",    "dj_addons"),
]

_BILLING_ROWS = [
    ("Billing Amount",  "billing_amount"),
    ("Advance Paid",    "advance_paid"),
    ("Pending Amount",  "pending_amount"),
    ("Coupon / Referral Code Applied", "coupon_code"),
]

# Manually-assigned fields — printed as blank lines for ops to fill by hand.
_MANUAL_ROWS = [
    "Event Ops Lead", "Decor Vendor", "Decor Reference Photo Confirmed",
    "Host Name", "DJ Name", "Entry Song", "Photographer Add-ons",
    "Cake", "Volunteers", "Piñata Bags", "Piñata Fillings",
    "Return Gift Wrapping (Y/N)", "Paper Bag (Y/N + size)",
    "Personalization (Y/N + design)", "Host Gifts (# confirmed)",
    "Advance Payment Mode", "Pending — Expected Payment Mode",
]


# ─── EXCEL (openpyxl) ─────────────────────────────────────────────────────

def _xl_image(image_bytes, width, height):
    if not image_bytes:
        return None
    try:
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(io.BytesIO(image_bytes))
        img.width, img.height = width, height
        return img
    except Exception as exc:
        logger.warning(f"order form: couldn't load image into xlsx — {exc}")
        return None


def build_order_form_xlsx(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Form"
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, size=16, color="FFFFFF")
    section_font = Font(bold=True, size=11, color="FFFFFF")
    label_font = Font(bold=True, size=10)
    value_font = Font(size=10)
    title_fill = PatternFill("solid", fgColor="7A4FBF")
    section_fill = PatternFill("solid", fgColor="B59DDE")
    manual_fill = PatternFill("solid", fgColor="FFF7E6")

    for col, width in zip("ABCDEFGHI", [22, 4, 26, 4, 22, 4, 26, 4, 4]):
        ws.column_dimensions[col].width = width

    # Title
    ws.merge_cells("A1:I2")
    c = ws["A1"]
    c.value = f"Wondershop Experiences — Birthday Order Form  (Lead #{data['lead_id']})"
    c.font = title_font
    c.fill = title_fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    left_row = 4
    right_row = 4

    def section(row, col_start, col_end, text):
        r = ws.cell(row=row, column=col_start)
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        r.value = text
        r.font = section_font
        r.fill = section_fill
        r.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 16
        return row + 1

    def kv(row, label_col, value_col, value_end_col, label, value, fill=None):
        lc = ws.cell(row=row, column=label_col, value=label)
        lc.font = label_font
        lc.border = border
        ws.merge_cells(start_row=row, start_column=value_col, end_row=row, end_column=value_end_col)
        vc = ws.cell(row=row, column=value_col, value=value)
        vc.font = value_font
        vc.border = border
        vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if fill:
            lc.fill = fill
            vc.fill = fill
        return row + 1

    # ── LEFT column: event details, services, billing ──
    left_row = section(left_row, 1, 3, "Event Details")
    for label, key in _DETAIL_ROWS:
        left_row = kv(left_row, 1, 3, 3, label, data.get(key, "—"))

    left_row += 1
    left_row = section(left_row, 1, 3, "Services (as booked)")
    for label, key in _SERVICE_ROWS:
        left_row = kv(left_row, 1, 3, 3, label, data.get(key, "—"))
    left_row = kv(left_row, 1, 3, 3, "Engagement Activities", ", ".join(data.get("activities", ["—"])))
    left_row = kv(left_row, 1, 3, 3, "Return Gifts", ", ".join(data.get("gifts", ["—"])))

    left_row += 1
    left_row = section(left_row, 1, 3, "Billing")
    for label, key in _BILLING_ROWS:
        left_row = kv(left_row, 1, 3, 3, label, data.get(key, "—"))

    left_row += 1
    left_row = section(left_row, 1, 3, "Remarks / Personalization Requests")
    ws.merge_cells(start_row=left_row, start_column=1, end_row=left_row + 1, end_column=3)
    rc = ws.cell(row=left_row, column=1, value=data.get("remarks", "—"))
    rc.font = value_font
    rc.border = border
    rc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    left_row += 3

    # ── RIGHT column: to-be-filled-by-ops, event schedule (page 1) ──
    # Inventory To Be Packed moves to its own page 3 (see below); reference
    # photos + return gifts move to page 2.
    right_row = section(right_row, 5, 9, "To Be Filled By Ops")
    for label in _MANUAL_ROWS:
        right_row = kv(right_row, 5, 7, 9, label, "", fill=manual_fill)

    right_row += 1
    right_row = section(right_row, 5, 9, "Event Schedule")
    ws.merge_cells(start_row=right_row, start_column=5, end_row=right_row, end_column=6)
    hc = ws.cell(row=right_row, column=5, value="Time")
    hc.font = label_font
    hc.border = border
    hc.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=right_row, start_column=7, end_row=right_row, end_column=9)
    hc2 = ws.cell(row=right_row, column=7, value="Activity")
    hc2.font = label_font
    hc2.border = border
    hc2.alignment = Alignment(horizontal="center")
    right_row += 1
    for _ in range(8):
        ws.merge_cells(start_row=right_row, start_column=5, end_row=right_row, end_column=6)
        ws.cell(row=right_row, column=5, value="").border = border
        ws.merge_cells(start_row=right_row, start_column=7, end_row=right_row, end_column=9)
        ws.cell(row=right_row, column=7, value="").border = border
        right_row += 1

    # Page break: Event Details / Ops / Billing / Remarks / Event Schedule
    # stay on page 1 — reference photos + return gifts start fresh on page 2.
    page1_end_row = max(left_row, right_row)
    ws.row_breaks.append(Break(id=page1_end_row))
    band_row = page1_end_row + 1

    # ── Decor reference photo + included/not-included spec ──
    decor_title = data.get("decor_name") or "—"
    band_row = section(band_row, 1, 9, f"Decor Reference — {decor_title}")
    img_row = band_row
    img = _xl_image(data.get("decor_image_bytes"), 190, 130)
    if img:
        ws.add_image(img, f"A{img_row}")
    spec = data.get("decor_spec") or []
    spec_row = band_row
    if spec:
        for label, value, na in spec:
            ws.merge_cells(start_row=spec_row, start_column=4, end_row=spec_row, end_column=5)
            lc = ws.cell(row=spec_row, column=4, value=label)
            lc.font = label_font
            lc.border = border
            ws.merge_cells(start_row=spec_row, start_column=6, end_row=spec_row, end_column=9)
            vc = ws.cell(row=spec_row, column=6, value=("✗ " + value if na else value))
            vc.font = Font(size=10, color="B00020" if na else "000000", italic=na)
            vc.border = border
            spec_row += 1
    else:
        ws.merge_cells(start_row=spec_row, start_column=4, end_row=spec_row, end_column=9)
        ws.cell(row=spec_row, column=4, value="No matching reference photo/inclusion list — Ops to confirm manually").font = value_font
        spec_row += 1
    band_row = max(spec_row, img_row + 9) + 1

    # ── Photographer inclusions ──
    photo_tier = data.get("photo_tier") or "—"
    band_row = section(band_row, 1, 9, f"Photographer — {photo_tier} Package Includes")
    features = data.get("photo_features") or ["—"]
    ws.merge_cells(start_row=band_row, start_column=1, end_row=band_row + max(1, len(features)) - 1, end_column=9)
    fc = ws.cell(row=band_row, column=1, value="\n".join(f"• {f}" for f in features))
    fc.font = value_font
    fc.border = border
    fc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    band_row += max(2, len(features)) + 1

    # ── E-Invite + Piñata reference photos ──
    band_row = section(band_row, 1, 9, "E-Invite Template & Piñata Reference")
    ws.cell(row=band_row, column=1, value="E-Invite:").font = label_font
    ws.cell(row=band_row, column=6, value="Piñata:").font = label_font
    ei_img = _xl_image(data.get("einvite_image_bytes"), 110, 150)
    pin_img = _xl_image(data.get("pinata_image_bytes"), 130, 130)
    if ei_img:
        ws.add_image(ei_img, f"A{band_row + 1}")
    else:
        ws.cell(row=band_row + 1, column=1, value="(no existing template chosen)").font = Font(size=9, italic=True, color="808080")
    if pin_img:
        ws.add_image(pin_img, f"F{band_row + 1}")
    else:
        ws.cell(row=band_row + 1, column=6, value="(custom design / not in catalogue)").font = Font(size=9, italic=True, color="808080")
    band_row += 10

    # ── Return Gifts — image, per-item value, packaging ──
    band_row = section(band_row, 1, 9, "Return Gifts")
    hdr = band_row
    for label, cs, ce in [("Photo", 1, 2), ("Item", 3, 5), ("Qty", 6, 6), ("Unit ₹", 7, 7), ("Total ₹", 8, 9)]:
        ws.merge_cells(start_row=hdr, start_column=cs, end_row=hdr, end_column=ce)
        hc = ws.cell(row=hdr, column=cs, value=label)
        hc.font = label_font
        hc.border = border
        hc.alignment = Alignment(horizontal="center")
    band_row += 1
    gifts_detail = data.get("gifts_detail") or []
    if not gifts_detail:
        ws.merge_cells(start_row=band_row, start_column=1, end_row=band_row, end_column=9)
        ws.cell(row=band_row, column=1, value="No return gifts on this booking").font = value_font
        band_row += 1
    for g in gifts_detail:
        ws.row_dimensions[band_row].height = 40
        ws.merge_cells(start_row=band_row, start_column=1, end_row=band_row, end_column=2)
        ws.cell(row=band_row, column=1).border = border
        gimg = _xl_image(g.get("image_bytes"), 45, 45)
        if gimg:
            ws.add_image(gimg, f"A{band_row}")
        ws.merge_cells(start_row=band_row, start_column=3, end_row=band_row, end_column=5)
        ic = ws.cell(row=band_row, column=3, value=g["name"])
        ic.font = value_font
        ic.border = border
        ic.alignment = Alignment(vertical="center", wrap_text=True)
        qc = ws.cell(row=band_row, column=6, value=g["qty"])
        qc.border = border
        qc.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=band_row, start_column=7, end_row=band_row, end_column=7)
        uc = ws.cell(row=band_row, column=7, value=(f"₹{g['unit']:,.0f}" if g.get("unit") is not None else "—"))
        uc.border = border
        uc.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=band_row, start_column=8, end_row=band_row, end_column=9)
        tc = ws.cell(row=band_row, column=8, value=(f"₹{g['total']:,.0f}" if g.get("total") is not None else "—"))
        tc.border = border
        tc.alignment = Alignment(horizontal="center", vertical="center")
        band_row += 1

    band_row = kv(band_row, 1, 3, 4, "Packaging", data.get("gift_packaging_label", "—"))
    ws.merge_cells(start_row=band_row - 1, start_column=6, end_row=band_row - 1, end_column=6)
    lc2 = ws.cell(row=band_row - 1, column=6, value="Personalized Thank You Note")
    lc2.font = label_font
    lc2.border = border
    ws.merge_cells(start_row=band_row - 1, start_column=7, end_row=band_row - 1, end_column=9)
    vc2 = ws.cell(row=band_row - 1, column=7, value=data.get("gift_thank_you_note", "—"))
    vc2.font = value_font
    vc2.border = border

    # Page break: reference photos + return gifts end here — inventory
    # checklist starts fresh on page 3.
    page2_end_row = band_row
    ws.row_breaks.append(Break(id=page2_end_row))
    band_row = page2_end_row + 2

    # ── Inventory To Be Packed (page 3) ──
    band_row = section(band_row, 1, 9, "Inventory To Be Packed")
    hdr_row = band_row
    headers = ["Type", "Item", "Qty", "Remarks"]
    cols = [1, 3, 7, 8]
    ends = [2, 6, 7, 9]
    for h, cs, ce in zip(headers, cols, ends):
        ws.merge_cells(start_row=hdr_row, start_column=cs, end_row=hdr_row, end_column=ce)
        hc = ws.cell(row=hdr_row, column=cs, value=h)
        hc.font = label_font
        hc.border = border
        hc.alignment = Alignment(horizontal="center")
    band_row += 1
    for _ in range(15):
        for cs, ce in zip(cols, ends):
            ws.merge_cells(start_row=band_row, start_column=cs, end_row=band_row, end_column=ce)
            ws.cell(row=band_row, column=cs, value="").border = border
        band_row += 1

    last_row = band_row + 1
    ws.row_dimensions[last_row].height = 8

    # Printer-friendly A4 landscape, fit to page width (content now spans
    # more than one printed page for bookings with several reference photos
    # — that's fine, each page is still clean and readable).
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4
    ws.print_area = f"A1:I{last_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── PDF (reportlab) ──────────────────────────────────────────────────────

_PURPLE = colors.HexColor("#7A4FBF")
_LIGHT_PURPLE = colors.HexColor("#B59DDE")
_MANUAL_BG = colors.HexColor("#FFF7E6")


def _kv_table(rows, col_widths, styles):
    body = getSampleStyleSheet()["BodyText"]
    body.fontSize = 8.5
    body.leading = 10.5
    data = []
    for label, value in rows:
        data.append([Paragraph(f"<b>{label}</b>", body), Paragraph(str(value), body)])
    t = Table(data, colWidths=col_widths, repeatRows=0)
    t.setStyle(TableStyle(styles + [
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    return t


def _section_header(text, width):
    styles = getSampleStyleSheet()
    st = ParagraphStyle("section", parent=styles["Normal"], textColor=colors.white,
                         fontSize=10, fontName="Helvetica-Bold", leftIndent=4, spaceBefore=0, spaceAfter=0)
    t = Table([[Paragraph(text, st)]], colWidths=[width])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _LIGHT_PURPLE),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return t


def build_order_form_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=8 * mm, bottomMargin=8 * mm,
        title="Birthday Order Form",
    )
    usable_w = page_size[0] - 20 * mm
    left_w = usable_w * 0.52
    right_w = usable_w * 0.44

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Normal"], textColor=colors.white,
                                  fontSize=14, fontName="Helvetica-Bold", leftIndent=4)
    title_tbl = Table([[Paragraph(f"Wondershop Experiences — Birthday Order Form (Lead #{data['lead_id']})", title_style)]],
                       colWidths=[usable_w])
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _PURPLE),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))

    left_flow = []
    left_flow.append(_section_header("Event Details", left_w))
    left_flow.append(_kv_table([(l, data.get(k, "—")) for l, k in _DETAIL_ROWS],
                                [left_w * 0.4, left_w * 0.6], []))
    left_flow.append(Spacer(1, 4))
    left_flow.append(_section_header("Services (as booked)", left_w))
    svc_rows = [(l, data.get(k, "—")) for l, k in _SERVICE_ROWS]
    svc_rows.append(("Engagement Activities", ", ".join(data.get("activities", ["—"]))))
    left_flow.append(_kv_table(svc_rows, [left_w * 0.4, left_w * 0.6], []))
    left_flow.append(Spacer(1, 4))
    left_flow.append(_section_header("Billing", left_w))
    left_flow.append(_kv_table([(l, data.get(k, "—")) for l, k in _BILLING_ROWS],
                                [left_w * 0.4, left_w * 0.6], []))
    left_flow.append(Spacer(1, 4))
    left_flow.append(_section_header("Remarks / Personalization Requests", left_w))
    body = styles["BodyText"]; body.fontSize = 8.5; body.leading = 10.5
    remarks_tbl = Table([[Paragraph(data.get("remarks", "—"), body)]], colWidths=[left_w])
    remarks_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 20),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]))
    left_flow.append(remarks_tbl)

    right_flow = []
    right_flow.append(_section_header("To Be Filled By Ops", right_w))
    manual_rows = [(label, "") for label in _MANUAL_ROWS]
    right_flow.append(_kv_table(manual_rows, [right_w * 0.55, right_w * 0.45],
                                 [("BACKGROUND", (1, 0), (1, -1), _MANUAL_BG),
                                  ("BOTTOMPADDING", (1, 0), (1, -1), 10)]))
    right_flow.append(Spacer(1, 4))
    right_flow.append(_section_header("Event Schedule", right_w))
    sched_header = [Paragraph("<b>Time</b>", body), Paragraph("<b>Activity</b>", body)]
    sched_data = [sched_header] + [["", ""] for _ in range(7)]
    sched_tbl = Table(sched_data, colWidths=[right_w * 0.25, right_w * 0.75])
    sched_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]))
    right_flow.append(sched_tbl)

    # Everything must fit on a single printed page — shrink-to-fit rather
    # than error if a booking has an unusually long remarks/services list.
    avail_h = page_size[1] - doc.topMargin - doc.bottomMargin - 110
    left_frame = KeepInFrame(left_w, avail_h, left_flow, mode="shrink")
    right_frame = KeepInFrame(right_w, avail_h, right_flow, mode="shrink")

    outer = Table([[left_frame, right_frame]], colWidths=[left_w, right_w])
    outer.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (1, 0), (1, 0), 8),
    ]))

    # Page 1: Event Details / Services / Billing / Remarks / Ops / Schedule.
    story = [title_tbl, Spacer(1, 6), outer]

    # Page 2: reference photos + return gifts.
    story.append(PageBreak())
    story.extend(_build_reference_bands(data, usable_w, styles, body))

    # Page 3: inventory checklist.
    story.append(PageBreak())
    story.append(_section_header("Inventory To Be Packed", usable_w))
    inv_header = [Paragraph(f"<b>{h}</b>", body) for h in ["Type", "Item", "Qty", "Remarks"]]
    inv_data = [inv_header] + [["", "", "", ""] for _ in range(15)]
    inv_tbl = Table(inv_data, colWidths=[usable_w * 0.18, usable_w * 0.37, usable_w * 0.12, usable_w * 0.33], repeatRows=1)
    inv_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(inv_tbl)

    doc.build(story)
    return buf.getvalue()


def _rl_image(image_bytes, max_w, max_h):
    """Aspect-ratio-preserving Image flowable from raw bytes, or None if
    there's nothing to show / the bytes can't be decoded."""
    if not image_bytes:
        return None
    try:
        from PIL import Image as PILImage
        from reportlab.platypus import Image as RLImage
        pil_img = PILImage.open(io.BytesIO(image_bytes))
        w, h = pil_img.size
        scale = min(max_w / w, max_h / h)
        return RLImage(io.BytesIO(image_bytes), width=w * scale, height=h * scale)
    except Exception as exc:
        logger.warning(f"order form: couldn't load image into pdf — {exc}")
        return None


def _build_reference_bands(data, usable_w, styles, body):
    """Decor reference photo + inclusions, Photographer inclusions,
    E-Invite/Piñata photos, and the Return Gifts table — flows naturally
    onto a second page if it doesn't fit under the details band above."""
    flow = []

    # ── Decor ──
    decor_title = data.get("decor_name") or "—"
    flow.append(_section_header(f"Decor Reference — {decor_title}", usable_w))
    img = _rl_image(data.get("decor_image_bytes"), usable_w * 0.28, 150)
    spec = data.get("decor_spec") or []
    if spec:
        spec_rows = []
        for label, value, na in spec:
            v = ("✗ " + value) if na else value
            spec_rows.append([Paragraph(f"<b>{label}</b>", body),
                               Paragraph(f'<font color="{"#B00020" if na else "#000000"}">{v}</font>', body)])
        spec_tbl = Table(spec_rows, colWidths=[usable_w * 0.28, usable_w * 0.44])
        spec_tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ]))
    else:
        spec_tbl = Paragraph("No matching reference photo/inclusion list — Ops to confirm manually", body)
    row = [[img if img else Paragraph("(no reference photo matched)", body), spec_tbl]]
    decor_row_tbl = Table(row, colWidths=[usable_w * 0.30, usable_w * 0.70])
    decor_row_tbl.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    flow.append(decor_row_tbl)
    flow.append(Spacer(1, 6))

    # ── Photographer ──
    photo_tier = data.get("photo_tier") or "—"
    flow.append(_section_header(f"Photographer — {photo_tier} Package Includes", usable_w))
    features = data.get("photo_features") or ["—"]
    feat_tbl = Table([[Paragraph("<br/>".join(f"• {f}" for f in features), body)]], colWidths=[usable_w])
    feat_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]))
    flow.append(feat_tbl)
    flow.append(Spacer(1, 6))

    # ── E-Invite + Piñata ──
    flow.append(_section_header("E-Invite Template & Piñata Reference", usable_w))
    ei_img = _rl_image(data.get("einvite_image_bytes"), usable_w * 0.20, 150)
    pin_img = _rl_image(data.get("pinata_image_bytes"), usable_w * 0.20, 150)
    ei_cell = ei_img if ei_img else Paragraph("(no existing template chosen)", body)
    pin_cell = pin_img if pin_img else Paragraph("(custom design / not in catalogue)", body)
    ei_pin_tbl = Table([[Paragraph("<b>E-Invite</b>", body), Paragraph("<b>Piñata</b>", body)],
                         [ei_cell, pin_cell]], colWidths=[usable_w * 0.5, usable_w * 0.5])
    ei_pin_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]))
    flow.append(ei_pin_tbl)
    flow.append(Spacer(1, 6))

    # ── Return Gifts ──
    flow.append(_section_header("Return Gifts", usable_w))
    gifts_detail = data.get("gifts_detail") or []
    gift_rows = [[Paragraph("<b>Photo</b>", body), Paragraph("<b>Item</b>", body),
                  Paragraph("<b>Qty</b>", body), Paragraph("<b>Unit (Rs.)</b>", body),
                  Paragraph("<b>Total (Rs.)</b>", body)]]
    if not gifts_detail:
        gift_rows.append([Paragraph("No return gifts on this booking", body), "", "", "", ""])
    for g in gifts_detail:
        thumb = _rl_image(g.get("image_bytes"), 40, 40) or Paragraph("—", body)
        unit_s = f"Rs. {g['unit']:,.0f}" if g.get("unit") is not None else "—"
        total_s = f"Rs. {g['total']:,.0f}" if g.get("total") is not None else "—"
        gift_rows.append([thumb, Paragraph(g["name"], body), Paragraph(str(g["qty"]), body),
                           Paragraph(unit_s, body), Paragraph(total_s, body)])
    gift_tbl = Table(gift_rows, colWidths=[usable_w * 0.10, usable_w * 0.50, usable_w * 0.12,
                                            usable_w * 0.14, usable_w * 0.14])
    gift_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F2F2")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]))
    flow.append(gift_tbl)
    flow.append(Spacer(1, 4))
    pack_tbl = Table([[Paragraph("<b>Packaging</b>", body), Paragraph(data.get("gift_packaging_label", "—"), body),
                        Paragraph("<b>Personalized Thank You Note</b>", body),
                        Paragraph(str(data.get("gift_thank_you_note", "—")), body)]],
                      colWidths=[usable_w * 0.15, usable_w * 0.35, usable_w * 0.25, usable_w * 0.25])
    pack_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]))
    flow.append(pack_tbl)

    return flow
